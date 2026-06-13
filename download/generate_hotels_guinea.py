import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.properties.creator = "Z.ai"

# ============ STYLES ============
header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
title_font = Font(name='Calibri', bold=True, size=16, color='1F4E79')
subtitle_font = Font(name='Calibri', bold=True, size=11, color='2E75B6')
data_font = Font(name='Calibri', size=10)
star_font = Font(name='Calibri', size=10, color='FF8C00')
city_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
alt_fill = PatternFill(start_color='F2F7FC', end_color='F2F7FC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0')
)

# ============ HOTEL DATA ============
# Comprehensive list compiled from: Booking.com, TripAdvisor, Expedia, Airbnb, Agoda,
# Hotels.com, Skyscanner, Petit Futé, Go Africa Online, Trivago, KAYAK, Trip.com,
# Wego, Cozycozy, HotelsCombined, Cybo, etc.

hotels = [
    # ============ CONAKRY ============
    # 5 étoiles
    {"nom": "Sheraton Grand Conakry", "ville": "Conakry", "quartier": "Kipe", "etoiles": 5, "plateformes": "Booking, Expedia, Agoda, TripAdvisor, Klook, Hotels.com, Trivago", "type": "Hôtel", "notes": "282 chambres, plage privée, piscine, centre de conférence. Premier hôtel Marriott en Guinée."},
    {"nom": "Radisson Blu Hotel Conakry", "ville": "Conakry", "quartier": "Kipé, Ratoma", "etoiles": 5, "plateformes": "Booking, Expedia, TripAdvisor, Agoda, Hotels.com, Trivago, KAYAK", "type": "Hôtel", "notes": "Piscine extérieure, vue océan, plus grand centre de conférence de Conakry. Note Expedia 9.2/10."},
    {"nom": "Noom Hotel Conakry", "ville": "Conakry", "quartier": "Camayenne", "etoiles": 5, "plateformes": "Booking, Expedia, TripAdvisor, Agoda, Hotels.com, Trivago, KAYAK, Wego", "type": "Hôtel", "notes": "Piscine saisonnière, WiFi gratuit. Groupe Mangalis. Note TripAdvisor 4.6. 766+ avis."},
    {"nom": "Palm Camayenne", "ville": "Conakry", "quartier": "Camayenne", "etoiles": 5, "plateformes": "Booking, Expedia, TripAdvisor, Agoda, Hotels.com, Trivago, KAYAK", "type": "Hôtel", "notes": "Plusieurs restaurants, connexion internet fiable, piscine. Considéré meilleur hôtel de Conakry par beaucoup."},
    {"nom": "Riviera Royal Hotel", "ville": "Conakry", "quartier": "Centre d'affaires", "etoiles": 5, "plateformes": "Booking, Expedia, TripAdvisor, Hotels.com, Trivago, Wego", "type": "Hôtel", "notes": "Piscine extérieure, quartier d'affaires. Note TripAdvisor 4.6. 466+ avis."},
    
    # 4 étoiles
    {"nom": "Millenium Suites", "ville": "Conakry", "quartier": "Landréah, Dixinn", "etoiles": 4, "plateformes": "Booking, TripAdvisor, YouTube, Wego, Petit Futé", "type": "Hôtel", "notes": "102 suites et 52 chambres, kitchenette, salon, piscine. Hôtel de standing."},
    {"nom": "Hôtel Le Petit Bateau", "ville": "Conakry", "quartier": "Corniche Nord, Kaloum", "etoiles": 4, "plateformes": "Booking, TripAdvisor, Expedia, Trip.com, Petit Futé, Planet of Hotels", "type": "Hôtel", "notes": "Piscine, petit-déjeuner gratuit, plage, navette aéroport gratuite. Note 7.9/10."},
    {"nom": "ONOMO Hotel Conakry", "ville": "Conakry", "quartier": "Camayenne", "etoiles": 4, "plateformes": "Booking, Expedia, TripAdvisor, Agoda, Hotels.com, Trivago, KAYAK, Petit Futé", "type": "Hôtel", "notes": "Note TripAdvisor #1 Conakry business. ATM dans le lobby. À partir de 66€/nuit."},
    {"nom": "Hôtel Azur Conakry", "ville": "Conakry", "quartier": "Camayenne", "etoiles": 4, "plateformes": "Booking, Expedia, Agoda, Cozycozy, Hotels.com", "type": "Hôtel", "notes": "Jardin, restaurant, bar. Note Agoda 6.8."},
    {"nom": "Hotel Kaloum", "ville": "Conakry", "quartier": "Kaloum", "etoiles": 4, "plateformes": "Booking, Expedia, TripAdvisor, Hotels.com, Trivago, HotelsOne, KAYAK", "type": "Hôtel", "notes": "Piscine, spa, salle de sport. Note Trip.com 'Very Good' 27 avis."},
    {"nom": "Souaré Premium Hôtel", "ville": "Conakry", "quartier": "Camayenne", "etoiles": 4, "plateformes": "Booking, Expedia, TripAdvisor, Hotels.com, Trivago, Wego, Skyscanner", "type": "Hôtel", "notes": "Hôtel premium récent. Note Skyscanner 2.53 (prix intérêt)."},
    {"nom": "Grand Hotel Central (GHC)", "ville": "Conakry", "quartier": "Kaloum", "etoiles": 4, "plateformes": "Booking, Expedia, TripAdvisor, Agoda, Hotels.com, Trivago, KAYAK, Petit Futé", "type": "Hôtel", "notes": "Hôtel historique au centre-ville. Note Agoda 6.8."},
    {"nom": "Riviera Taouyah Hotel", "ville": "Conakry", "quartier": "Taouyah", "etoiles": 3, "plateformes": "Booking, TripAdvisor, Expedia, Hotels.com, Skyscanner, Petit Futé", "type": "Hôtel", "notes": "Centre de fitness, jardin, restaurant, bar. WiFi gratuit. 3 étoiles."},
    {"nom": "HOTEL GOLDEN PLAZZA CONAKRY", "ville": "Conakry", "quartier": "Kaloum", "etoiles": 3, "plateformes": "Booking, Expedia, Agoda, TripAdvisor, Hotels.com, Trivago, Traveloka, Momondo", "type": "Hôtel", "notes": "Note Expedia 10/10. Chambres propres et spacieuses. 7 min de l'ambassade de France. À partir de 60€."},
    {"nom": "Zambezi Inn Airport Hotel", "ville": "Conakry", "quartier": "Près aéroport", "etoiles": 4, "plateformes": "Booking, Expedia, TripAdvisor, Agoda, Hotels.com, Trivago, KAYAK, Skyscanner", "type": "Hôtel", "notes": "Proche aéroport international. Note Agoda 'très bien' propreté. À partir de 66€/nuit."},
    {"nom": "Atlantic View Hotel", "ville": "Conakry", "quartier": "Corniche", "etoiles": 4, "plateformes": "Booking, Expedia, TripAdvisor, Hotels.com, Trivago, KAYAK", "type": "Hôtel", "notes": "Vue sur l'Atlantique. Toutes chambres avec vue mer."},
    {"nom": "Hotel Mariador Park", "ville": "Conakry", "quartier": "Camayenne", "etoiles": 4, "plateformes": "Booking, TripAdvisor, Transform Africa Summit", "type": "Hôtel", "notes": "Hôtel de conférence. Partenaire événements."},
    {"nom": "Hôtel du Golfe de Guinée", "ville": "Conakry", "quartier": "Résidentiel", "etoiles": 3, "plateformes": "Booking, Agoda, Petit Futé, Cozycozy, ZenHotels, TravelMyth", "type": "Hôtel", "notes": "3 étoiles, vue sur le golfe, terrasse, parking gratuit, restaurant. Quartier résidentiel calme."},
    {"nom": "Hôtel M'lys", "ville": "Conakry", "quartier": "Camayenne", "etoiles": 3, "plateformes": "Cozycozy, Petit Futé", "type": "Hôtel", "notes": "Hôtel populaire mentionné par Cozycozy parmi les 2-3 étoiles."},
    {"nom": "Le Petit Chalet Hotel", "ville": "Conakry", "quartier": "Conakry", "etoiles": 3, "plateformes": "TripAdvisor", "type": "Hôtel", "notes": "30 chambres. Toutes taxes incluses."},
    {"nom": "Hakaba", "ville": "Conakry", "quartier": "Près aéroport", "etoiles": 3, "plateformes": "Expedia, Trip.com, Booking", "type": "Hôtel", "notes": "Proche aéroport. Option économique."},
    {"nom": "Grand Hôtel de l'Indépendance", "ville": "Conakry", "quartier": "Kaloum", "etoiles": 3, "plateformes": "TripAdvisor, Hotels.com, Travel Weekly", "type": "Hôtel", "notes": "Hôtel historique, ancien. Avis mitigés sur TripAdvisor."},
    {"nom": "Barry-Tall Guest House", "ville": "Conakry", "quartier": "Conakry", "etoiles": 2, "plateformes": "Momondo", "type": "Guest House", "notes": "Maison d'hôtes. Option économique."},
    {"nom": "Studio meublé à Conakry", "ville": "Conakry", "quartier": "Conakry", "etoiles": 0, "plateformes": "Agoda", "type": "Appartement", "notes": "Studio meublé. Très bien noté pour la propreté sur Agoda."},
    {"nom": "Tahi Residence - Conakry 01", "ville": "Conakry", "quartier": "Conakry", "etoiles": 0, "plateformes": "Airbnb", "type": "Condominium", "notes": "Appartement entier, 2 chambres, 2 salles de bain. Note Airbnb 4.78."},
    {"nom": "La Petite Minière (Auberge)", "ville": "Conakry", "quartier": "Minière, Dixinn", "etoiles": 1, "plateformes": "Petit Futé", "type": "Auberge", "notes": "Auberge économique quartier Minière."},
    {"nom": "Hotel Restaurant Océano", "ville": "Conakry", "quartier": "Conakry", "etoiles": 3, "plateformes": "TravelMyth", "type": "Hôtel", "notes": "Note TravelMyth 8.6. Restaurant sur place."},
    {"nom": "Climax Hotels", "ville": "Conakry", "quartier": "Conakry", "etoiles": 3, "plateformes": "Go Africa Online", "type": "Hôtel", "notes": "Hôtel, bar, restaurant, piscine. Tél: (+224) 669 00 09 99."},
    {"nom": "Hôtel de l'Université", "ville": "Conakry", "quartier": "Conakry", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné par Petit Futé."},
    {"nom": "Le Mariador Linsan", "ville": "Conakry", "quartier": "Linsan", "etoiles": 3, "plateformes": "TripAdvisor", "type": "Hôtel", "notes": "Hôtel spécialisé. TripAdvisor Kindia Region."},
    {"nom": "Gospel Voice", "ville": "Conakry", "quartier": "Conakry", "etoiles": 1, "plateformes": "TripAdvisor", "type": "Chambre d'hôtes", "notes": "Chambre d'hôtes économique."},
    {"nom": "Setifana Sea View", "ville": "Conakry", "quartier": "Corniche", "etoiles": 3, "plateformes": "Hotels.com, Hoteles.com", "type": "Hôtel", "notes": "Vue sur mer."},
    
    # ============ KINDIA ============
    {"nom": "Hotel Kanya", "ville": "Kindia", "quartier": "Kindia", "etoiles": 3, "plateformes": "TripAdvisor, Booking, Hotels.com", "type": "Hôtel", "notes": "Note TripAdvisor 4.3. Chambres climatisées, TV, eau chaude, moustiquaire. Très bien noté."},
    {"nom": "Les Eaux de Kilissi", "ville": "Kindia", "quartier": "Kilissi", "etoiles": 3, "plateformes": "TripAdvisor", "type": "Hôtel", "notes": "Note TripAdvisor 4.0. 6 avis. Source thermale."},
    {"nom": "Hôtel Masabi", "ville": "Kindia", "quartier": "Kindia", "etoiles": 3, "plateformes": "Petit Futé, Site web hotelmasabi.com", "type": "Hôtel", "notes": "Chambres de 380 000 GNF (standard) à 600 000 GNF (deluxe). Vue jardin."},
    {"nom": "Hôtel Moringa", "ville": "Kindia", "quartier": "Kindia", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné par Petit Futé Kindia."},
    {"nom": "Hôtel Phare de Guinée", "ville": "Kindia", "quartier": "Kindia", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné par Petit Futé Kindia."},
    {"nom": "La Cabane Bambou", "ville": "Kindia", "quartier": "Kindia", "etoiles": 1, "plateformes": "Petit Futé", "type": "Auberge", "notes": "Hébergement économique Petit Futé."},
    {"nom": "Hôtel Sooli (Chaîne Mangrove)", "ville": "Kindia", "quartier": "Kindia", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Chaîne Mangrove. Mentionné Petit Futé."},
    {"nom": "Hôtel Buffet de la Gare", "ville": "Kindia", "quartier": "Kindia", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Proche gare. Mentionné Petit Futé."},
    {"nom": "Hôtel Marcica", "ville": "Kindia", "quartier": "Kindia", "etoiles": 3, "plateformes": "Facebook", "type": "Hôtel", "notes": "33 chambres, restaurant gastronomique, salle de réunion."},
    
    # ============ KANKAN ============
    {"nom": "Hôtel Bâté", "ville": "Kankan", "quartier": "Kankan", "etoiles": 3, "plateformes": "Petit Futé, Cybo", "type": "Hôtel", "notes": "Mentionné par Petit Futé et Cybo parmi les meilleurs de Kankan."},
    {"nom": "Hôtel Résidence de la Ville de Kankan (RVK)", "ville": "Kankan", "quartier": "Kankan", "etoiles": 3, "plateformes": "Facebook", "type": "Résidence", "notes": "Appartements meublés et climatisés. 285 likes Facebook."},
    {"nom": "Hôtel Résidence Nabaya", "ville": "Kankan", "quartier": "Kankan", "etoiles": 2, "plateformes": "Petit Futé, Cybo", "type": "Hôtel", "notes": "Mentionné Petit Futé Kankan."},
    {"nom": "Hôtel Plazza Senkéfra", "ville": "Kankan", "quartier": "Senkéfra", "etoiles": 2, "plateformes": "Cybo", "type": "Hôtel", "notes": "Mentionné par Cybo."},
    {"nom": "Le Calao", "ville": "Kankan", "quartier": "Kankan", "etoiles": 2, "plateformes": "Cybo", "type": "Hôtel", "notes": "Mentionné par Cybo Kankan."},
    {"nom": "Hôtel Lamartine", "ville": "Kankan", "quartier": "Kankan", "etoiles": 2, "plateformes": "Cybo", "type": "Hôtel", "notes": "Mentionné par Cybo Kankan."},
    {"nom": "Villa Sylli", "ville": "Kankan", "quartier": "Kankan", "etoiles": 2, "plateformes": "Cybo", "type": "Hôtel", "notes": "Mentionné par Cybo Kankan."},
    {"nom": "Hôtel Philadelphia (ex-Relais Bâté)", "ville": "Kankan", "quartier": "Kankan", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Anciennement Relais Bâté. Mentionné Petit Futé."},
    
    # ============ NZÉRÉKORÉ ============
    {"nom": "Hôtel le Prince", "ville": "Nzérékoré", "quartier": "Nzérékoré", "etoiles": 3, "plateformes": "TripAdvisor", "type": "Hôtel", "notes": "#1 TripAdvisor Nzérékoré."},
    {"nom": "Hôtel Mont Nimba", "ville": "Nzérékoré", "quartier": "Ossid", "etoiles": 3, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Quartier Ossid. Mentionné Petit Futé."},
    {"nom": "Hôtel Diani", "ville": "Nzérékoré", "quartier": "Nzérékoré", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné Petit Futé Nzérékoré."},
    {"nom": "Hôtel Le Palmier", "ville": "Nzérékoré", "quartier": "Nien", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Quartier Nien. Mentionné Petit Futé."},
    {"nom": "Hôtel Golo", "ville": "Nzérékoré", "quartier": "Tilépoulou", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Quartier Tilépoulou. Mentionné Petit Futé."},
    {"nom": "Hôtel Ivoire", "ville": "Nzérékoré", "quartier": "Centre", "etoiles": 2, "plateformes": "LocaHotels", "type": "Hôtel", "notes": "Adresse centrale, pratique pour rayonner à pied."},
    {"nom": "Hôtel Nimba", "ville": "Nzérékoré", "quartier": "Nzérékoré", "etoiles": 2, "plateformes": "LocaHotels", "type": "Hôtel", "notes": "Base confortable pour excursions vers Lola & Bossou."},
    {"nom": "Hôtel la Forêt", "ville": "Nzérékoré", "quartier": "Nzérékoré", "etoiles": 2, "plateformes": "LocaHotels", "type": "Hôtel", "notes": "Mentionné par LocaHotels Nzérékoré."},
    {"nom": "Zaly Transit Motel", "ville": "Nzérékoré", "quartier": "Nzérékoré", "etoiles": 1, "plateformes": "Zaly Merveille", "type": "Motel", "notes": "Quartier calme, chambres propres et confortables."},
    {"nom": "Beer Garden Hotel, Bar & Restaurant", "ville": "Nzérékoré", "quartier": "Nzérékoré", "etoiles": 2, "plateformes": "TripAdvisor", "type": "Hôtel", "notes": "Mentionné TripAdvisor Nzérékoré."},
    {"nom": "Nimba Ecolodge & Reserve", "ville": "Nzérékoré", "quartier": "Monts Nimba", "etoiles": 4, "plateformes": "TripAdvisor, Hotels.com, Site web nimbaecolodge.com", "type": "Écolodge", "notes": "Note TripAdvisor 5.0. 11 cabanes éco-responsables. Piscine, canoë, randonnées. 220$/nuit. 51 km du centre."},

    # ============ LABÉ ============
    {"nom": "Hôtel Tata", "ville": "Labé", "quartier": "Labé", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné Petit Futé Labé."},
    {"nom": "Hôtel Salaa+", "ville": "Labé", "quartier": "Labé", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné Petit Futé Labé."},
    {"nom": "Complexe Alizée Provincial", "ville": "Labé", "quartier": "Labé", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné Petit Futé Labé."},
    {"nom": "Hôtel Safatou", "ville": "Labé", "quartier": "Labé", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné Petit Futé Labé."},
    {"nom": "Motel Riviera", "ville": "Labé", "quartier": "Labé", "etoiles": 1, "plateformes": "Petit Futé", "type": "Motel", "notes": "Mentionné Petit Futé Labé."},
    {"nom": "Djamtum Hotel", "ville": "Labé", "quartier": "Labé", "etoiles": 2, "plateformes": "Facebook, Site web djamtum.com", "type": "Hôtel", "notes": "Site web djamtum.com. Tél: +224 620 02 55 30."},
    {"nom": "Résidence Universitaire de Labé", "ville": "Labé", "quartier": "Labé", "etoiles": 1, "plateformes": "Petit Futé", "type": "Résidence", "notes": "Mentionné Petit Futé Labé."},
    
    # ============ MAMOU ============
    {"nom": "Hôtel Baly's", "ville": "Mamou", "quartier": "Mamou", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Chambres climatisées 350 000-500 000 FG, ventilées 300 000 FG. Petit-déjeuner inclus."},
    {"nom": "Hotel Safitel Dalaba", "ville": "Mamou", "quartier": "Dalaba", "etoiles": 2, "plateformes": "TripAdvisor", "type": "Hôtel", "notes": "#1 TripAdvisor Mamou Region. Note 3.0."},
    {"nom": "Lodge Kamawi", "ville": "Mamou", "quartier": "Mamou", "etoiles": 2, "plateformes": "TripAdvisor", "type": "Lodge", "notes": "#2 TripAdvisor Mamou Region. Note 4.0."},
    
    # ============ FARANAH ============
    {"nom": "Hôtel Del Niger", "ville": "Faranah", "quartier": "Faranah", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné Petit Futé Faranah. Également listé pour Kankan."},
    {"nom": "Hôtel Firya", "ville": "Faranah", "quartier": "Faranah", "etoiles": 2, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné Petit Futé Faranah."},
    {"nom": "Hôtel Bibish", "ville": "Faranah", "quartier": "Faranah", "etoiles": 1, "plateformes": "Petit Futé, Zaly Merveille", "type": "Hôtel", "notes": "Option simple et économique, bon rapport qualité-prix."},
    {"nom": "Hôtel Sandenya", "ville": "Faranah", "quartier": "Faranah", "etoiles": 1, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné Petit Futé Faranah."},
    {"nom": "Hôtel Silibe", "ville": "Faranah", "quartier": "Faranah", "etoiles": 1, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné Petit Futé Faranah."},
    
    # ============ SIGUIRI ============
    {"nom": "Hôtel Relais Bâté (Siguiri)", "ville": "Siguiri", "quartier": "Siguiri", "etoiles": 1, "plateformes": "Casa Trotter, Go Africa Online", "type": "Hôtel", "notes": "Plats 30 000-50 000 FG. Tél: 66 25 26 97."},
    {"nom": "Délices d'Éden", "ville": "Siguiri", "quartier": "Siguirinkoua II", "etoiles": 1, "plateformes": "Go Africa Online", "type": "Hôtel", "notes": "Boulegard, 2ème Station. Mentionné Go Africa Online."},
    
    # ============ DABOLA / KOUROUSSA ============
    {"nom": "Hôtel Tinkisso", "ville": "Dabola", "quartier": "Dabola", "etoiles": 1, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Mentionné Petit Futé (alentours de Kankan)."},
    {"nom": "Hôtel Tando", "ville": "Kouroussa", "quartier": "Quartier Raiko", "etoiles": 1, "plateformes": "Petit Futé", "type": "Hôtel", "notes": "Quartier Raiko, Kouroussa. Mentionné Petit Futé."},
    
    # ============ BOKÉ / KAMSAR ============
    {"nom": "Hôtel La Corniche Kamsar", "ville": "Kamsar", "quartier": "Kamsar", "etoiles": 3, "plateformes": "Hotels.com", "type": "Hôtel", "notes": "Mentionné Hotels.com Boké Region."},
    {"nom": "Setifana Sea View", "ville": "Kamsar", "quartier": "Kamsar", "etoiles": 2, "plateformes": "Hotels.com", "type": "Hôtel", "notes": "Vue mer. Mentionné Hotels.com Boké."},
    
    # ============ DUBRÉKA ============
    {"nom": "Souaré Club Hôtel", "ville": "Dubréka", "quartier": "Dubréka", "etoiles": 3, "plateformes": "Booking, TripAdvisor", "type": "Hôtel", "notes": "#1 TripAdvisor Dubréka. Hôtel spécialité."},
    {"nom": "Résidence Diakhaby", "ville": "Dubréka", "quartier": "Dubréka", "etoiles": 2, "plateformes": "Booking", "type": "Résidence", "notes": "Mentionné Booking.com Dubréka."},
    {"nom": "Résidences MAFCA", "ville": "Dubréka", "quartier": "Dubréka", "etoiles": 2, "plateformes": "Booking", "type": "Résidence", "notes": "Mentionné Booking.com Dubréka."},
    
    # ============ FRIA ============
    {"nom": "Hôtel de Fria", "ville": "Fria", "quartier": "Fria", "etoiles": 2, "plateformes": "Booking", "type": "Hôtel", "notes": "Disponible sur Booking.com Fria."},
    
    # ============ FORÉCARIAH ============
    {"nom": "Hôtel de Forécariah", "ville": "Forécariah", "quartier": "Forécariah", "etoiles": 1, "plateformes": "MakeMyTrip", "type": "Hôtel", "notes": "Mentionné MakeMyTrip liste villes Guinée."},
    
    # ============ AUTRES (Sources diverses) ============
    {"nom": "Centre d'Accueil Diocésain", "ville": "Kankan", "quartier": "Kankan", "etoiles": 1, "plateformes": "Cybo", "type": "Centre d'accueil", "notes": "Mentionné Cybo Kankan."},
    {"nom": "Konakri (Hôtel)", "ville": "Kankan", "quartier": "Kankan", "etoiles": 1, "plateformes": "Cybo", "type": "Hôtel", "notes": "Mentionné Cybo Kankan."},
    {"nom": "Buffet de la Gare (Kankan)", "ville": "Kankan", "quartier": "Kankan", "etoiles": 1, "plateformes": "Cybo", "type": "Hôtel", "notes": "Mentionné Cybo Kankan."},
]

# ============ SHEET 1: LISTE COMPLETE ============
ws1 = wb.active
ws1.title = "Liste complète"

# Title
ws1.merge_cells('A1:H1')
ws1['A1'] = "RÉPERTOIRE COMPLET DES HÔTELS EN GUINÉE"
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('A2:H2')
ws1['A2'] = f"Données compilées de Booking.com, TripAdvisor, Expedia, Airbnb, Agoda, Hotels.com, Skyscanner, Petit Futé, Go Africa Online, Trivago, KAYAK, Trip.com, Wego, Cybo et plus — {len(hotels)} hôtels répertoriés"
ws1['A2'].font = Font(name='Calibri', size=9, italic=True, color='666666')
ws1['A2'].alignment = Alignment(horizontal='center')

# Headers
headers = ["#", "Nom de l'hôtel", "Ville", "Quartier / Localisation", "Catégorie (étoiles)", "Type", "Plateformes sources", "Notes / Détails"]
for col_idx, header in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Data
current_city = None
for idx, hotel in enumerate(hotels, 1):
    row = idx + 4
    
    # City separator row
    if hotel["ville"] != current_city:
        current_city = hotel["ville"]
        # We'll just use alternating colors instead of city separator for cleaner look
    
    ws1.cell(row=row, column=1, value=idx).font = data_font
    ws1.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    
    ws1.cell(row=row, column=2, value=hotel["nom"]).font = Font(name='Calibri', size=10, bold=True)
    ws1.cell(row=row, column=3, value=hotel["ville"]).font = data_font
    ws1.cell(row=row, column=4, value=hotel["quartier"]).font = data_font
    ws1.cell(row=row, column=5, value="★" * hotel["etoiles"] if hotel["etoiles"] > 0 else "Non classé").font = star_font
    ws1.cell(row=row, column=5).alignment = Alignment(horizontal='center')
    ws1.cell(row=row, column=6, value=hotel["type"]).font = data_font
    ws1.cell(row=row, column=7, value=hotel["plateformes"]).font = Font(name='Calibri', size=9, color='2E75B6')
    ws1.cell(row=row, column=8, value=hotel["notes"]).font = Font(name='Calibri', size=9, color='555555')
    ws1.cell(row=row, column=8).alignment = Alignment(wrap_text=True)
    
    # Alternating row colors
    if idx % 2 == 0:
        for col in range(1, 9):
            ws1.cell(row=row, column=col).fill = alt_fill
    
    # Borders
    for col in range(1, 9):
        ws1.cell(row=row, column=col).border = thin_border

# Column widths
col_widths = [5, 42, 16, 22, 18, 16, 40, 50]
for i, width in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# Freeze panes
ws1.freeze_panes = 'A5'

# Auto filter
ws1.auto_filter.ref = f"A4:H{len(hotels)+4}"

# ============ SHEET 2: PAR VILLE ============
ws2 = wb.create_sheet("Par ville")

ws2.merge_cells('A1:F1')
ws2['A1'] = "HÔTELS PAR VILLE"
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')

# Group by city
from collections import OrderedDict
cities = OrderedDict()
for h in hotels:
    city = h["ville"]
    if city not in cities:
        cities[city] = []
    cities[city].append(h)

row = 3
for city, city_hotels in cities.items():
    # City header
    ws2.merge_cells(f'A{row}:F{row}')
    ws2.cell(row=row, column=1, value=f"{city} ({len(city_hotels)} hôtel{'s' if len(city_hotels)>1 else ''})")
    ws2.cell(row=row, column=1).font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    ws2.cell(row=row, column=1).fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    ws2.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    row += 1
    
    # Sub-headers
    sub_headers = ["Nom", "Quartier", "Étoiles", "Type", "Plateformes", "Notes"]
    for col_idx, header in enumerate(sub_headers, 1):
        cell = ws2.cell(row=row, column=col_idx, value=header)
        cell.font = Font(name='Calibri', bold=True, size=10, color='1F4E79')
        cell.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        cell.border = thin_border
    row += 1
    
    for h in city_hotels:
        ws2.cell(row=row, column=1, value=h["nom"]).font = Font(name='Calibri', size=10, bold=True)
        ws2.cell(row=row, column=2, value=h["quartier"]).font = data_font
        ws2.cell(row=row, column=3, value="★" * h["etoiles"] if h["etoiles"] > 0 else "NC").font = star_font
        ws2.cell(row=row, column=4, value=h["type"]).font = data_font
        ws2.cell(row=row, column=5, value=h["plateformes"]).font = Font(name='Calibri', size=9, color='2E75B6')
        ws2.cell(row=row, column=6, value=h["notes"]).font = Font(name='Calibri', size=9, color='555555')
        ws2.cell(row=row, column=6).alignment = Alignment(wrap_text=True)
        for col in range(1, 7):
            ws2.cell(row=row, column=col).border = thin_border
        row += 1
    row += 1  # Space between cities

col_widths2 = [42, 22, 14, 16, 40, 50]
for i, width in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# ============ SHEET 3: STATISTIQUES ============
ws3 = wb.create_sheet("Statistiques")

ws3.merge_cells('A1:D1')
ws3['A1'] = "STATISTIQUES HÔTELS EN GUINÉE"
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center')

# Stats by city
row = 3
ws3.cell(row=row, column=1, value="Nombre d'hôtels par ville").font = Font(name='Calibri', bold=True, size=13, color='1F4E79')
row += 1
for col_idx, header in enumerate(["Ville", "Nombre d'hôtels", "5 étoiles", "4 étoiles"], 1):
    cell = ws3.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

row += 1
for city, city_hotels in cities.items():
    five = sum(1 for h in city_hotels if h["etoiles"] == 5)
    four = sum(1 for h in city_hotels if h["etoiles"] == 4)
    ws3.cell(row=row, column=1, value=city).font = Font(name='Calibri', size=10, bold=True)
    ws3.cell(row=row, column=2, value=len(city_hotels)).font = data_font
    ws3.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws3.cell(row=row, column=3, value=five).font = data_font
    ws3.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    ws3.cell(row=row, column=4, value=four).font = data_font
    ws3.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    for col in range(1, 5):
        ws3.cell(row=row, column=col).border = thin_border
    row += 1

# Total
row += 1
ws3.cell(row=row, column=1, value="TOTAL").font = Font(name='Calibri', bold=True, size=11)
ws3.cell(row=row, column=2, value=len(hotels)).font = Font(name='Calibri', bold=True, size=11)
ws3.cell(row=row, column=2).alignment = Alignment(horizontal='center')
ws3.cell(row=row, column=3, value=sum(1 for h in hotels if h["etoiles"]==5)).font = Font(name='Calibri', bold=True, size=11)
ws3.cell(row=row, column=3).alignment = Alignment(horizontal='center')
ws3.cell(row=row, column=4, value=sum(1 for h in hotels if h["etoiles"]==4)).font = Font(name='Calibri', bold=True, size=11)
ws3.cell(row=row, column=4).alignment = Alignment(horizontal='center')
for col in range(1, 5):
    ws3.cell(row=row, column=col).border = thin_border

# Stats by category
row += 3
ws3.cell(row=row, column=1, value="Répartition par catégorie").font = Font(name='Calibri', bold=True, size=13, color='1F4E79')
row += 1
for col_idx, header in enumerate(["Catégorie", "Nombre"], 1):
    cell = ws3.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

categories = {}
for h in hotels:
    cat = f"{h['etoiles']} étoiles" if h["etoiles"] > 0 else "Non classé"
    categories[cat] = categories.get(cat, 0) + 1

row += 1
for cat, count in sorted(categories.items(), key=lambda x: x[1], reverse=True):
    ws3.cell(row=row, column=1, value=cat).font = data_font
    ws3.cell(row=row, column=2, value=count).font = data_font
    ws3.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    for col in range(1, 3):
        ws3.cell(row=row, column=col).border = thin_border
    row += 1

# Stats by type
row += 2
ws3.cell(row=row, column=1, value="Répartition par type d'hébergement").font = Font(name='Calibri', bold=True, size=13, color='1F4E79')
row += 1
for col_idx, header in enumerate(["Type", "Nombre"], 1):
    cell = ws3.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

types = {}
for h in hotels:
    t = h["type"]
    types[t] = types.get(t, 0) + 1

row += 1
for t, count in sorted(types.items(), key=lambda x: x[1], reverse=True):
    ws3.cell(row=row, column=1, value=t).font = data_font
    ws3.cell(row=row, column=2, value=count).font = data_font
    ws3.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    for col in range(1, 3):
        ws3.cell(row=row, column=col).border = thin_border
    row += 1

# Platform stats
row += 2
ws3.cell(row=row, column=1, value="Couverture par plateforme (nombre d'hôtels mentionnés)").font = Font(name='Calibri', bold=True, size=13, color='1F4E79')
row += 1
for col_idx, header in enumerate(["Plateforme", "Hôtels référencés"], 1):
    cell = ws3.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

platforms = {}
for h in hotels:
    for p in h["plateformes"].split(","):
        p = p.strip()
        platforms[p] = platforms.get(p, 0) + 1

row += 1
for p, count in sorted(platforms.items(), key=lambda x: x[1], reverse=True):
    ws3.cell(row=row, column=1, value=p).font = data_font
    ws3.cell(row=row, column=2, value=count).font = data_font
    ws3.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    for col in range(1, 3):
        ws3.cell(row=row, column=col).border = thin_border
    row += 1

# Column widths for stats
ws3.column_dimensions['A'].width = 42
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15

# ============ SHEET 4: SOURCES ============
ws4 = wb.create_sheet("Sources")

ws4.merge_cells('A1:C1')
ws4['A1'] = "SOURCES DES DONNÉES"
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center')

sources = [
    ("Booking.com", "booking.com", "47 hôtels Conakry, villes secondaires (Kindia, Nzérékoré, Kankan, Labé, Fria, Dubréka, Kamsar)"),
    ("TripAdvisor", "tripadvisor.com", "Avis et classements Conakry, Kindia, Nzérékoré, Mamou, Labé, Faranah, Dubréka"),
    ("Expedia", "expedia.com", "Hôtels Guinée toutes villes. Notes détaillées."),
    ("Airbnb", "airbnb.com", "Locations vacances Conakry et environs. Appartements et condominiums."),
    ("Agoda", "agoda.com", "11+ hôtels Conakry. Notes propreté détaillées."),
    ("Hotels.com / Hoteles.com", "hotels.com", "Classements Conakry, Kindia, Boké, Dubréka"),
    ("Skyscanner", "skyscanner.net", "Comparaison prix hôtels Conakry, Siguiri, Kindia"),
    ("Trivago", "trivago.com", "142 hôtels Guinée, 133 à Conakry"),
    ("KAYAK / HotelsCombined", "kayak.fr / hotelscombined.com", "Comparaison prix 48+ hôtels Conakry"),
    ("Petit Futé", "petitfute.com", "Guide touristique complet : Conakry, Kindia, Kankan, Nzérékoré, Labé, Mamou, Faranah, Siguiri"),
    ("Go Africa Online", "goafricaonline.com", "Annuaire hôtels Guinée. Contacts directs."),
    ("Trip.com", "trip.com", "Hôtels Conakry, Kankan, Nzérékoré, Labé, Faranah, Kindia"),
    ("Wego", "wego.com", "47+ hôtels Conakry. Comparaison prix."),
    ("Momondo", "momondo.com / momondo.fr", "48 hôtels Conakry. Golden Plazza, Petit Bateau, etc."),
    ("Cozycozy", "cozycozy.com", "Comparaison hôtels et locations Guinée"),
    ("Cybo", "cybo.com", "Hôtels Kankan : Bâté, Plazza Senkéfra, Nabaya, etc."),
    ("LocaHotels", "locahotels.com", "Guide hôtels Nzérékoré"),
    ("Casa Trotter", "casa-trotter.com", "Hôtels Siguiri et nord-est Guinée"),
    ("Traveloka", "traveloka.com", "Hôtels Kankan Region, Sangarédi"),
    ("MakeMyTrip", "makemytrip.global", "Liste villes : Conakry, Boké, Forécariah"),
    ("Zaly Merveille", "zalymerveille.com", "Hôtels Nzérékoré, Faranah"),
    ("Transform Africa Summit", "transformafricasummit.org", "Hôtels partenaires événements Conakry"),
    "divider",
    ("Source officielle", "iaprp.org (PDF)", "Liste officielle des hôtels de Guinée avec tarifs négociés"),
]

row = 3
for col_idx, header in enumerate(["Plateforme / Source", "Site web", "Couverture"], 1):
    cell = ws4.cell(row=row, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

row = 4
for source in sources:
    if source == "divider":
        row += 1
        continue
    ws4.cell(row=row, column=1, value=source[0]).font = Font(name='Calibri', size=10, bold=True)
    ws4.cell(row=row, column=2, value=source[1]).font = Font(name='Calibri', size=9, color='2E75B6')
    ws4.cell(row=row, column=3, value=source[2]).font = Font(name='Calibri', size=9, color='555555')
    ws4.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
    for col in range(1, 4):
        ws4.cell(row=row, column=col).border = thin_border
    row += 1

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 60

# Save
output_path = "/home/z/my-project/download/Hotels_Guinee_Repertoire_Complet.xlsx"
wb.save(output_path)
print(f"Fichier sauvegardé : {output_path}")
print(f"Total hôtels : {len(hotels)}")
print(f"Villes couvertes : {len(cities)}")
